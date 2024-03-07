# -*- coding: utf-8 -*-
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.gaussian_process import GaussianProcessRegressor
from sklearn.gaussian_process.kernels import RBF
from sklearn.metrics import mean_squared_error
from scipy.integrate import trapz
plt.rcParams['font.sans-serif'] = ['SimHei']  # 避免中文出现乱码
plt.rcParams['axes.unicode_minus'] = False

# 读取Excel数据
df = pd.read_excel(r"D:\新建文件夹 (7)\shuwien1.xlsx")
x = df['宽度'].tolist()
y = df['高度'].tolist()
z = 3.54
x = [x for x in x if x is not None and not pd.isna(x)]
y = [x for x in y if x is not None and not pd.isna(x)]
X = np.array(x).reshape(-1, 1)
y = np.array(y)


# 定义高斯过程回归模型
kernel = RBF(length_scale=1.0, length_scale_bounds=(1e-1, 10.0))
gpr = GaussianProcessRegressor(kernel=kernel, alpha=0.1, n_restarts_optimizer=10)
# 训练模型
gpr.fit(X, y)
# 预测新数据
X_new = np.linspace(0, 35, 100).reshape(-1, 1)
y_pred, y_std = gpr.predict(X_new, return_std=True)
# 计算均方误差
y_true = np.interp(x, X_new.flatten(), y_pred.flatten())
mse = mean_squared_error(y, y_true)
print("均方误差（MSE）：", mse)
# 绘制图形

plt.figure(figsize=(12,4),dpi=400)
plt.plot(X_new, y_pred, 'b-', label='预测')
plt.fill_between(X_new[:,0], y_pred - y_std, y_pred + y_std, alpha=0.2, color='blue')
plt.plot(x, y, 'r-', marker='o', markersize=5, label='已有')
plt.xticks(range(0, 35, 1))
plt.title('河水断面剖面图')
plt.xlabel('宽度/m')
plt.ylabel('高度/m')
plt.legend(loc='upper left')
plt.gca().invert_yaxis()
plt.axhline(y=z, color='gray', linestyle='--')
plt.show()
# 提取拟合曲线高度大于3.54的部分
indices = np.where(y_pred > z)[0]
X_new_selected = X_new[indices]
y_pred_selected = y_pred[indices]
X_new_list = X_new_selected.flatten().tolist()
y_pred_list = y_pred_selected.flatten().tolist()
y_minus_3_54 = np.subtract(y_pred_list, 3.54)
# 计算数值积分
integral_value = trapz(y_minus_3_54, X_new_list)
print("对给定数据进行数值积分的结果为:", integral_value)
print(y_minus_3_54)

end_y = max(y_minus_3_54)  # 起始 y 值
step_size = 0.01  # y 的步长
num_steps = 10  # 循环次数，即计算 10 个积分值
y_values = np.linspace(0, end_y, num_steps)

integral_values = []  # 用于存储积分值的列表
depth=[]
#积分方法对于每一次的水深y值列表，和x的求梯形积分。
for i in y_values:
    y_minus_c = np.subtract(y_minus_3_54, i)
    y_minus_c_non_negative = np.maximum(y_minus_c, 0)# 计算 y - c，剔除负值变为0
    integral = trapz(y_minus_c_non_negative, X_new_list)  # 计算数值积分
    integral_values.append(integral)
    depth.append(end_y-i)# 将积分值添加到列表中
    print(f"水深为 {end_y-i} 时的积分值为 {integral}")
    if  integral<0:
        break
    else:
        continue

print("所有y值对应的积分值:", integral_values)
plt.figure(figsize=(12,4),dpi=400)  # 创建一个新的图形窗口
plt.plot( integral_values,depth,'b-', marker='.')
plt.title('拟合水深——面积图')
plt.xlabel('面积/m*m')
plt.ylabel('水深/m')
plt.show()


import openpyxl



# 将列表转换为Numpy数组
array1 = np.array(depth)
array2 = np.array(integral_values)

# 打开指定的工作簿
workbook = openpyxl.load_workbook(r"D:\新建文件夹 (7)\shuwien1.xlsx")

# 选择默认的活动工作表
worksheet = workbook.active

# 将Numpy数组的值写入Excel的两列
for i in range(len(array1)):
    worksheet.cell(row=i+2, column=7, value=array1[i])
    worksheet.cell(row=i+2, column=8, value=array2[i])

# 保存工作簿到文件
workbook.save(r"D:\新建文件夹 (7)\shuwien1.xlsx")
